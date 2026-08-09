"""
Excel processing: append POC Status and highlight FALSE/NA cells.

Flow:
  1. Load workbook from storage/uploads (.xlsx directly, or convert .xls first).
  2. Add a new column at the end with header POC Status.
  3. For each data row (non-summary, has content), set Processed and red-fill FALSE/NA
     in one configured column (default Match_DOB).
  4. Add worksheet ``Name and SSN`` with columns name and ssn copied from the source sheet.
  5. Save to storage/processed/processed_<uuid>.xlsx.

Uses a single index pass over worksheet cells (avoids O(rows × cells) scans).
"""

import logging
import os
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from app.services.excel.excel_config import (
    EXTRACT_SHEET_NAME,
    EXTRACT_TAB_NAME_HEADER,
    EXTRACT_TAB_SSN_HEADER,
    FALSE_NA_COLUMN_HEADERS,
    FALSE_NA_FILL_RGB,
    HEADER_ROW,
    NEW_COLUMN_DEFAULT_VALUE,
    NEW_COLUMN_HEADER,
    SOURCE_NAME_HEADERS,
    SOURCE_SSN_HEADERS,
)
from app.services.excel.xls_converter import convert_xls_to_xlsx

logger = logging.getLogger(__name__)

_RED_FILL = PatternFill(fill_type="solid", fgColor=FALSE_NA_FILL_RGB)


class ExcelProcessor:
    """
    Transforms an uploaded .xlsx or legacy .xls into a processed .xlsx copy with an extra status column.

    Summary rows (containing 'total' in columns A–E) are skipped for POC values.
    """

    def process(self, source_path: Path, output_path: Path) -> dict[str, str | int]:
        """
        Run the full Excel pipeline and write ``output_path``.

        Args:
            source_path: Path to the uploaded .xlsx or .xls in storage/uploads.
            output_path: Destination .xlsx path under storage/processed.

        Returns:
            Metadata for the API response (sheet name, column index, counts).
        """
        if not source_path.is_file():
            raise FileNotFoundError(f"Source file not found: {source_path}")

        logger.info("Processing Excel file: %s", source_path.name)

        temp_dir: tempfile.TemporaryDirectory[str] | None = None
        work_path = source_path
        if source_path.suffix.lower() == ".xls":
            temp_dir = tempfile.TemporaryDirectory(prefix="poc_xls_")
            work_path = Path(temp_dir.name) / f"{source_path.stem}.xlsx"
            convert_xls_to_xlsx(source_path, work_path)

        try:
            cached_false_na = self._load_false_na_column_cache(work_path)
            workbook = load_workbook(work_path, data_only=False)
            worksheet = workbook.active
            result = self._apply_transforms(workbook, worksheet, cached_false_na)
            self._mark_workbook_for_recalc(workbook)

            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            workbook.close()

            logger.info(
                "Saved processed file: %s (rows updated: %s, new column: %s, red fills: %s)",
                output_path.name,
                result["rows_updated"],
                result["new_column_index"],
                result["cells_filled_red"],
            )

            return result
        finally:
            if temp_dir is not None:
                temp_dir.cleanup()

    def process_inplace(self, target_path: Path) -> dict[str, str | int]:
        """
        Apply the same transforms as ``process`` and overwrite ``target_path`` in place.

        Uses a temporary file in the same directory, then atomic replace, to reduce
        risk of a partial write if save fails.
        """
        if not target_path.is_file():
            raise FileNotFoundError(f"Target file not found: {target_path}")
        if target_path.suffix.lower() != ".xlsx":
            raise ValueError("In-place processing supports .xlsx only")

        logger.info("In-place Excel update: %s", target_path)

        cached_false_na = self._load_false_na_column_cache(target_path)
        workbook = load_workbook(target_path, data_only=False)
        try:
            worksheet = workbook.active
            result = self._apply_transforms(workbook, worksheet, cached_false_na)
            self._mark_workbook_for_recalc(workbook)

            fd, tmp_name = tempfile.mkstemp(
                suffix=".xlsx",
                prefix=f".{target_path.stem}_poc_",
                dir=target_path.parent,
            )
            os.close(fd)
            tmp_path = Path(tmp_name)
            try:
                workbook.save(tmp_path)
                os.replace(tmp_path, target_path)
            finally:
                if tmp_path.is_file():
                    tmp_path.unlink(missing_ok=True)

            logger.info(
                "Updated in place: %s (rows updated: %s, column: %s)",
                target_path.name,
                result["rows_updated"],
                result["new_column_index"],
            )
            return result
        finally:
            workbook.close()

    def _mark_workbook_for_recalc(self, workbook) -> None:
        """Tell Excel to recalculate all formulas when the file is opened."""
        calc = workbook.calculation
        calc.calcMode = "auto"
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcCompleted = False
        calc.calcOnSave = False
        calc.calcId = 0

    def _load_false_na_column_cache(self, path: Path) -> dict[int, object]:
        """
        Stream cached calculated values for the FALSE/NA column (read_only + data_only).

        Faster than a second full in-memory workbook load for large sheets.
        """
        cache: dict[int, object] = {}
        try:
            values_wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            logger.exception("Could not load data_only cache from %s", path.name)
            return cache
        try:
            ws = values_wb.active
            header_row = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW), None)
            if header_row is None:
                return cache
            needles = {name.strip().lower() for name in FALSE_NA_COLUMN_HEADERS}
            col_idx = None
            for idx, cell in enumerate(header_row, start=1):
                if cell.value is None:
                    continue
                if str(cell.value).strip().lower() in needles:
                    col_idx = idx
                    break
            if col_idx is None:
                return cache
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=HEADER_ROW + 1, min_col=col_idx, max_col=col_idx),
                start=HEADER_ROW + 1,
            ):
                cache[row_idx] = row[0].value
            logger.info(
                "Cached %s FALSE/NA values for column %s (read_only)",
                len(cache),
                FALSE_NA_COLUMN_HEADERS,
            )
            return cache
        finally:
            values_wb.close()

    def _index_worksheet(self, worksheet) -> tuple[int, set[int], set[int]]:
        """
        One pass over existing cells.

        Returns:
            (last_data_row, rows_with_content, summary_rows)
        """
        last_row = HEADER_ROW
        content_rows: set[int] = set()
        summary_rows: set[int] = set()
        for (row, col), cell in worksheet._cells.items():
            if row <= HEADER_ROW:
                continue
            if self._cell_has_meaningful_content(cell):
                last_row = max(last_row, row)
                content_rows.add(row)
            if col <= 5 and cell.value is not None:
                if "total" in str(cell.value).strip().lower():
                    summary_rows.add(row)
        return max(last_row, HEADER_ROW + 1), content_rows, summary_rows

    def _apply_transforms(
        self,
        workbook,
        worksheet,
        cached_false_na: dict[int, object] | None = None,
    ) -> dict[str, str | int]:
        """POC Status column, FALSE/NA highlighting (one column), and Name/SSN extract sheet."""
        new_col = self._resolve_or_create_poc_column(worksheet)
        false_na_col = self._find_header_column(worksheet, FALSE_NA_COLUMN_HEADERS)
        if false_na_col is None:
            logger.warning(
                "FALSE/NA highlight column not found (expected one of %s); skipping fills",
                FALSE_NA_COLUMN_HEADERS,
            )

        last_data_row, content_rows, summary_rows = self._index_worksheet(worksheet)

        rows_updated = 0
        cells_filled_red = 0
        for row in range(HEADER_ROW + 1, last_data_row + 1):
            if row in summary_rows or row not in content_rows:
                continue
            worksheet.cell(row=row, column=new_col, value=NEW_COLUMN_DEFAULT_VALUE)
            if false_na_col is not None:
                check_value = None
                if cached_false_na is not None and row in cached_false_na:
                    check_value = cached_false_na[row]
                cells_filled_red += self._fill_false_na_in_column(
                    worksheet, row, false_na_col, check_value=check_value
                )
            rows_updated += 1

        name_ssn_rows = self._create_name_ssn_sheet(
            workbook,
            worksheet,
            last_data_row=last_data_row,
            summary_rows=summary_rows,
        )
        sheet_name = worksheet.title

        return {
            "sheet_name": sheet_name,
            "new_column_index": new_col,
            "new_column_header": NEW_COLUMN_HEADER,
            "rows_updated": rows_updated,
            "cells_filled_red": cells_filled_red,
            "extract_sheet_name": EXTRACT_SHEET_NAME,
            "name_ssn_rows": name_ssn_rows,
        }

    def _resolve_or_create_poc_column(self, worksheet) -> int:
        """Use existing POC Status column if present; otherwise append a new column."""
        for col in range(1, worksheet.max_column + 1):
            header = self._get_cell_value(worksheet, HEADER_ROW, col)
            if header is None:
                continue
            if str(header).strip() == NEW_COLUMN_HEADER:
                return col
        new_col = worksheet.max_column + 1
        worksheet.cell(row=HEADER_ROW, column=new_col, value=NEW_COLUMN_HEADER)
        return new_col

    def _fill_false_na_in_column(
        self,
        worksheet,
        row: int,
        col: int,
        check_value=None,
    ) -> int:
        """Apply red fill on ``row`` only in ``col`` when that cell is FALSE or NA."""
        cell = worksheet._cells.get((row, col))
        if cell is None:
            if check_value is None or not self._is_false_or_na_value(check_value):
                return 0
            cell = worksheet.cell(row=row, column=col)
        value = check_value if check_value is not None else cell.value
        if not self._is_false_or_na_value(value):
            return 0
        cell.fill = copy(_RED_FILL)
        return 1

    def _is_false_or_na_value(self, value) -> bool:
        """True for boolean False or string false / na / n/a (case-insensitive)."""
        if isinstance(value, bool):
            return value is False
        if value is None:
            return False
        text = str(value).strip().lower()
        return text in {"false", "na", "n/a"}

    def _cell_has_meaningful_content(self, cell) -> bool:
        """True if the cell has a formula or non-blank scalar value."""
        if cell.data_type == "f" and cell.value:
            return True
        if cell.value is None:
            return False
        return str(cell.value).strip() != ""

    def _get_cell_value(self, worksheet, row: int, col: int):
        """Read cell value without creating a new cell if missing."""
        cell = worksheet._cells.get((row, col))
        if cell is None:
            return None
        return cell.value

    def _find_header_column(self, worksheet, header_names: tuple[str, ...]) -> int | None:
        """Return 1-based column index whose row-1 header matches one of ``header_names``."""
        needles = {name.strip().lower() for name in header_names}
        for col in range(1, worksheet.max_column + 1):
            value = self._get_cell_value(worksheet, HEADER_ROW, col)
            if value is None:
                continue
            if str(value).strip().lower() in needles:
                return col
        return None

    def _create_name_ssn_sheet(
        self,
        workbook,
        worksheet,
        last_data_row: int | None = None,
        summary_rows: set[int] | None = None,
    ) -> int:
        """
        Add a tab with columns ``name`` and ``ssn`` populated from the active sheet.

        Skips summary rows and rows without a Name value.
        """
        name_col = self._find_header_column(worksheet, SOURCE_NAME_HEADERS)
        if name_col is None:
            raise ValueError(
                f'Name column not found in row {HEADER_ROW} (expected header: Name)'
            )
        ssn_col = self._find_header_column(worksheet, SOURCE_SSN_HEADERS)

        if EXTRACT_SHEET_NAME in workbook.sheetnames:
            del workbook[EXTRACT_SHEET_NAME]

        extract_ws = workbook.create_sheet(title=EXTRACT_SHEET_NAME[:31])
        extract_ws.cell(row=HEADER_ROW, column=1, value=EXTRACT_TAB_NAME_HEADER)
        extract_ws.cell(row=HEADER_ROW, column=2, value=EXTRACT_TAB_SSN_HEADER)

        if last_data_row is None or summary_rows is None:
            last_data_row, _content, summary_rows = self._index_worksheet(worksheet)

        out_row = HEADER_ROW + 1
        for row in range(HEADER_ROW + 1, last_data_row + 1):
            if row in summary_rows:
                continue
            name_val = self._get_cell_value(worksheet, row, name_col)
            if name_val is None or str(name_val).strip() == "":
                continue
            ssn_val = self._get_cell_value(worksheet, row, ssn_col) if ssn_col else None
            extract_ws.cell(row=out_row, column=1, value=name_val)
            extract_ws.cell(row=out_row, column=2, value=ssn_val)
            out_row += 1

        rows_written = out_row - HEADER_ROW - 1
        logger.info(
            "Created sheet %r with %s name/ssn rows (name col=%s, ssn col=%s)",
            EXTRACT_SHEET_NAME,
            rows_written,
            name_col,
            ssn_col,
        )
        return rows_written
