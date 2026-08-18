"""
Convert legacy Excel 97–2003 (.xls) workbooks to .xlsx for openpyxl processing.

openpyxl does not read .xls; xlrd extracts cell values (formatting is not preserved).
"""

from __future__ import annotations

import logging
from pathlib import Path

import xlrd
from openpyxl import Workbook

logger = logging.getLogger(__name__)


def convert_xls_to_xlsx(xls_path: Path, xlsx_path: Path) -> None:
    """
    Copy all sheets from ``xls_path`` into a new ``xlsx_path`` workbook.

    Args:
        xls_path: Source .xls file on disk.
        xlsx_path: Destination .xlsx path (parent directory must exist or be creatable).

    Raises:
        FileNotFoundError: source missing.
        xlrd.XLRDError: corrupt or unsupported workbook.
    """
    if not xls_path.is_file():
        raise FileNotFoundError(f"XLS file not found: {xls_path}")

    xls_book = xlrd.open_workbook(str(xls_path), formatting_info=False)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_index in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        title = (xls_sheet.name or f"Sheet{sheet_index + 1}")[:31]
        worksheet = workbook.create_sheet(title=title)

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell = xls_sheet.cell(row_idx, col_idx)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, xls_book.datemode)
                worksheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(xlsx_path)
    logger.info("Converted %s → %s (%s sheets)", xls_path.name, xlsx_path.name, xls_book.nsheets)
